import xmltodict
import os
import pandas as pd
import logging
import sys
import time

# Configuração de logging
logging.basicConfig(
    level=logging.DEBUG,  # Alterado para DEBUG para ver mais informações
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("processamento_xml.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

def extrair_dados(nota_xml, dados_coletados):
    try:
        logging.info(f"Iniciando processamento do arquivo {nota_xml}")
        with open(f'nfs/{nota_xml}', "rb") as xml_file:
            logging.info(f"Arquivo {nota_xml} aberto, iniciando parse XML")
            xml_dict = xmltodict.parse(xml_file)
            logging.info(f"Parse XML concluído para {nota_xml}")

            # Tenta encontrar a estrutura correta do XML
            if "NFe" in xml_dict:
                detalhes_nf = xml_dict["NFe"]['infNFe']
            elif "nfeProc" in xml_dict and "NFe" in xml_dict["nfeProc"]:
                detalhes_nf = xml_dict["nfeProc"]["NFe"]['infNFe']
            elif "ConsultarNfseResposta" in xml_dict:
                # Estrutura para NFSe (Nota Fiscal de Serviço Eletrônica)
                logging.info(f"Arquivo {nota_xml} é uma NFSe, estrutura diferente de NFe")
                detalhes_nf = xml_dict["ConsultarNfseResposta"]["ListaNfse"]["CompNfse"]["Nfse"]["InfNfse"]
                # Adaptar extração para NFSe
                numero = detalhes_nf.get("Numero", "Não informado")
                empresa = detalhes_nf.get("PrestadorServico", {}).get("RazaoSocial", "Não informado")
                cliente = detalhes_nf.get("TomadorServico", {}).get("RazaoSocial", "Não informado")
                endereco_cliente = "Não disponível em NFSe"
                peso_total = "Não aplicável a NFSe"

                dados_coletados.append([numero, empresa, cliente, endereco_cliente, peso_total])
                return
            elif "ConsultarNfseServicoPrestadoResposta" in xml_dict:
                # Estrutura para NFSe de Serviço Prestado
                logging.info(f"Arquivo {nota_xml} é uma NFSe de Serviço Prestado")

                # Verifica se há múltiplas notas no arquivo
                list_nfse = xml_dict["ConsultarNfseServicoPrestadoResposta"]["ListaNfse"]["CompNfse"]

                # Se for uma lista de notas, processa cada uma
                if isinstance(list_nfse, list):
                    total_notas = len(list_nfse)
                    logging.info(f"Arquivo {nota_xml} contém {total_notas} notas fiscais")

                    # Processa todas as notas com indicador de progresso
                    notas_processadas = 0
                    inicio = time.time()

                    # Pergunta ao usuário se deseja processar todas as notas
                    print(f"O arquivo contém {total_notas} notas fiscais. Processando todas as notas...")

                    for i, comp_nfse in enumerate(list_nfse):

                        # Mostra progresso a cada 500 notas (ou a cada 1% para arquivos muito grandes)
                        intervalo_progresso = max(500, int(total_notas * 0.01))
                        if i % intervalo_progresso == 0 and i > 0:
                            tempo_decorrido = time.time() - inicio
                            percentual = i/total_notas*100

                            # Calcula tempo estimado para conclusão
                            if tempo_decorrido > 0:
                                notas_por_segundo = i / tempo_decorrido
                                notas_restantes = total_notas - i
                                tempo_restante = notas_restantes / notas_por_segundo if notas_por_segundo > 0 else 0

                                # Formata o tempo restante
                                if tempo_restante < 60:
                                    tempo_restante_str = f"{tempo_restante:.0f} segundos"
                                elif tempo_restante < 3600:
                                    tempo_restante_str = f"{tempo_restante/60:.1f} minutos"
                                else:
                                    tempo_restante_str = f"{tempo_restante/3600:.1f} horas"

                                mensagem = f"Processadas {i} de {total_notas} notas ({percentual:.1f}%) - Tempo decorrido: {tempo_decorrido:.1f}s - Tempo restante estimado: {tempo_restante_str}"
                            else:
                                mensagem = f"Processadas {i} de {total_notas} notas ({percentual:.1f}%) - Tempo decorrido: {tempo_decorrido:.1f}s"

                            logging.info(mensagem)
                            print(mensagem)

                        try:
                            detalhes_nf = comp_nfse["Nfse"]["InfNfse"]

                            # Informações básicas da nota
                            numero = detalhes_nf.get("Numero", "Não informado")
                            data_emissao = detalhes_nf.get("DataEmissao", "Não informado")
                            codigo_verificacao = detalhes_nf.get("CodigoVerificacao", "Não informado")

                            # Informações do prestador
                            prestador = detalhes_nf.get("PrestadorServico", {})
                            empresa = prestador.get("RazaoSocial", "Não informado")

                            # CNPJ do prestador
                            cnpj_prestador = "Não informado"
                            if "IdentificacaoPrestador" in prestador:
                                id_prestador = prestador["IdentificacaoPrestador"]
                                if "CpfCnpj" in id_prestador:
                                    cpf_cnpj = id_prestador["CpfCnpj"]
                                    cnpj_prestador = cpf_cnpj.get("Cnpj", cpf_cnpj.get("Cpf", "Não informado"))

                            # Informações do município do prestador
                            municipio = "Não informado"
                            uf = "Não informado"
                            if "Endereco" in prestador:
                                endereco_prestador = prestador["Endereco"]
                                municipio = endereco_prestador.get("CodigoMunicipio", "Não informado")
                                uf = endereco_prestador.get("Uf", "Não informado")

                            # Informações da declaração de serviço
                            competencia = "Não informado"
                            valor_servicos = "0.00"
                            valor_iss = "0.00"
                            aliquota = "0.00"
                            discriminacao = "Não informado"

                            if "DeclaracaoPrestacaoServico" in detalhes_nf:
                                decl = detalhes_nf["DeclaracaoPrestacaoServico"]
                                if "InfDeclaracaoPrestacaoServico" in decl:
                                    inf_decl = decl["InfDeclaracaoPrestacaoServico"]
                                    competencia = inf_decl.get("Competencia", "Não informado")

                                    # Informações do serviço
                                    if "Servico" in inf_decl:
                                        servico = inf_decl["Servico"]
                                        discriminacao = servico.get("Discriminacao", "Não informado")

                                        # Valores do serviço
                                        if "Valores" in servico:
                                            valores = servico["Valores"]
                                            valor_servicos = valores.get("ValorServicos", "0.00")
                                            valor_iss = valores.get("ValorIss", "0.00")
                                            aliquota = valores.get("Aliquota", "0.00")

                            # Também verifica valores na raiz da nota
                            if valor_servicos == "0.00" and "ValoresNfse" in detalhes_nf:
                                valores_nfse = detalhes_nf["ValoresNfse"]
                                valor_servicos = valores_nfse.get("ValorLiquidoNfse", valor_servicos)
                                valor_iss = valores_nfse.get("ValorIss", valor_iss)

                            # Informações do tomador
                            tomador = None
                            if "DeclaracaoPrestacaoServico" in detalhes_nf:
                                tomador = detalhes_nf.get("DeclaracaoPrestacaoServico", {}).get("InfDeclaracaoPrestacaoServico", {}).get("Tomador", {})
                            else:
                                tomador = detalhes_nf.get("Tomador", {})

                            cliente = tomador.get("RazaoSocial", "Não informado") if tomador else "Não informado"

                            # CNPJ/CPF do tomador
                            cnpj_cpf_tomador = "Não informado"
                            if tomador and "IdentificacaoTomador" in tomador:
                                id_tomador = tomador["IdentificacaoTomador"]
                                if "CpfCnpj" in id_tomador:
                                    cpf_cnpj = id_tomador["CpfCnpj"]
                                    cnpj_cpf_tomador = cpf_cnpj.get("Cnpj", cpf_cnpj.get("Cpf", "Não informado"))

                            # Extrai campos adicionais
                            inscricao_municipal_prestador = "Não informado"
                            if "IdentificacaoPrestador" in prestador:
                                inscricao_municipal_prestador = prestador["IdentificacaoPrestador"].get("InscricaoMunicipal", "Não informado")

                            # Valores adicionais
                            base_calculo = "0.00"
                            if "ValoresNfse" in detalhes_nf:
                                valores_nfse = detalhes_nf["ValoresNfse"]
                                base_calculo = valores_nfse.get("BaseCalculo", valores_nfse.get("ValorServicos", "0.00"))
                                logging.debug(f"ValoresNfse encontrado: {valores_nfse}")

                            # Informações de ISS
                            iss_retido = "Não"
                            municipio_incidencia = "Não informado"
                            optante_simples = "Não"
                            incentivo_fiscal = "Não"

                            # Códigos de serviço
                            item_lista_servico = "Não informado"
                            codigo_cnae = "Não informado"
                            codigo_tributacao = "Não informado"

                            if "DeclaracaoPrestacaoServico" in detalhes_nf:
                                decl = detalhes_nf["DeclaracaoPrestacaoServico"]
                                if "InfDeclaracaoPrestacaoServico" in decl:
                                    inf_decl = decl["InfDeclaracaoPrestacaoServico"]

                                    # Informações fiscais
                                    optante_simples = "Sim" if inf_decl.get("OptanteSimplesNacional", "0") == "1" else "Não"
                                    incentivo_fiscal = "Sim" if inf_decl.get("IncentivoFiscal", "0") == "1" else "Não"

                                    if "Servico" in inf_decl:
                                        servico = inf_decl["Servico"]

                                        # Códigos de serviço
                                        item_lista_servico = servico.get("ItemListaServico", "Não informado")
                                        codigo_cnae = servico.get("CodigoCnae", "Não informado")
                                        codigo_tributacao = servico.get("CodigoTributacaoMunicipio", "Não informado")

                                        # ISS
                                        iss_retido_codigo = servico.get("IssRetido", "0")
                                        iss_retido = "Sim" if iss_retido_codigo == "1" else "Não"

                                        # Município de incidência
                                        municipio_incidencia = servico.get("MunicipioIncidencia", "Não informado")

                            # Calcula o valor líquido (Base de Cálculo - Valor ISS)
                            try:
                                # Imprime os valores para debug
                                logging.debug(f"Base de Cálculo: {base_calculo}, tipo: {type(base_calculo)}")
                                logging.debug(f"Valor ISS: {valor_iss}, tipo: {type(valor_iss)}")

                                # Converte para float, tratando diferentes formatos
                                if isinstance(base_calculo, str):
                                    base_calculo_num = float(base_calculo.replace(',', '.'))
                                else:
                                    base_calculo_num = float(base_calculo) if base_calculo else 0.0

                                if isinstance(valor_iss, str):
                                    valor_iss_num = float(valor_iss.replace(',', '.'))
                                else:
                                    valor_iss_num = float(valor_iss) if valor_iss else 0.0

                                # Calcula o valor líquido
                                valor_liquido = base_calculo_num - valor_iss_num
                                valor_liquido = f"{valor_liquido:.2f}".replace('.', ',')

                                logging.debug(f"Valor Líquido calculado: {valor_liquido}")
                            except (ValueError, TypeError) as e:
                                logging.warning(f"Erro ao calcular valor líquido: {str(e)}")
                                logging.warning(f"Base de Cálculo: {base_calculo}, Valor ISS: {valor_iss}")
                                valor_liquido = "0,00"

                            # Adiciona todos os dados coletados
                            dados_coletados.append([
                                numero,
                                data_emissao,
                                competencia,
                                codigo_verificacao,
                                empresa,
                                cnpj_prestador,
                                inscricao_municipal_prestador,
                                cliente,
                                cnpj_cpf_tomador,
                                valor_servicos,
                                base_calculo,
                                aliquota,
                                valor_iss,
                                valor_liquido,
                                iss_retido,
                                item_lista_servico,
                                codigo_cnae,
                                codigo_tributacao,
                                discriminacao,
                                municipio,
                                uf,
                                municipio_incidencia,
                                optante_simples,
                                incentivo_fiscal
                            ])
                            notas_processadas += 1
                        except Exception as e:
                            logging.warning(f"Erro ao processar nota {i+1} do arquivo {nota_xml}: {str(e)}")

                    tempo_total = time.time() - inicio

                    # Formata o tempo total
                    if tempo_total < 60:
                        tempo_total_str = f"{tempo_total:.1f} segundos"
                    elif tempo_total < 3600:
                        tempo_total_str = f"{tempo_total/60:.1f} minutos"
                    else:
                        tempo_total_str = f"{tempo_total/3600:.2f} horas"

                    mensagem_final = f"Processamento concluído. {notas_processadas} notas processadas em {tempo_total_str}"
                    logging.info(mensagem_final)
                    print(mensagem_final)
                    return
                else:
                    # Se for apenas uma nota
                    try:
                        detalhes_nf = list_nfse["Nfse"]["InfNfse"]

                        # Informações básicas da nota
                        numero = detalhes_nf.get("Numero", "Não informado")
                        data_emissao = detalhes_nf.get("DataEmissao", "Não informado")
                        codigo_verificacao = detalhes_nf.get("CodigoVerificacao", "Não informado")

                        # Informações do prestador
                        prestador = detalhes_nf.get("PrestadorServico", {})
                        empresa = prestador.get("RazaoSocial", "Não informado")

                        # CNPJ do prestador
                        cnpj_prestador = "Não informado"
                        if "IdentificacaoPrestador" in prestador:
                            id_prestador = prestador["IdentificacaoPrestador"]
                            if "CpfCnpj" in id_prestador:
                                cpf_cnpj = id_prestador["CpfCnpj"]
                                cnpj_prestador = cpf_cnpj.get("Cnpj", cpf_cnpj.get("Cpf", "Não informado"))

                        # Informações do município do prestador
                        municipio = "Não informado"
                        uf = "Não informado"
                        if "Endereco" in prestador:
                            endereco_prestador = prestador["Endereco"]
                            municipio = endereco_prestador.get("CodigoMunicipio", "Não informado")
                            uf = endereco_prestador.get("Uf", "Não informado")

                        # Informações da declaração de serviço
                        competencia = "Não informado"
                        valor_servicos = "0.00"
                        valor_iss = "0.00"
                        aliquota = "0.00"
                        discriminacao = "Não informado"

                        if "DeclaracaoPrestacaoServico" in detalhes_nf:
                            decl = detalhes_nf["DeclaracaoPrestacaoServico"]
                            if "InfDeclaracaoPrestacaoServico" in decl:
                                inf_decl = decl["InfDeclaracaoPrestacaoServico"]
                                competencia = inf_decl.get("Competencia", "Não informado")

                                # Informações do serviço
                                if "Servico" in inf_decl:
                                    servico = inf_decl["Servico"]
                                    discriminacao = servico.get("Discriminacao", "Não informado")

                                    # Valores do serviço
                                    if "Valores" in servico:
                                        valores = servico["Valores"]
                                        valor_servicos = valores.get("ValorServicos", "0.00")
                                        valor_iss = valores.get("ValorIss", "0.00")
                                        aliquota = valores.get("Aliquota", "0.00")

                        # Também verifica valores na raiz da nota
                        if valor_servicos == "0.00" and "ValoresNfse" in detalhes_nf:
                            valores_nfse = detalhes_nf["ValoresNfse"]
                            valor_servicos = valores_nfse.get("ValorLiquidoNfse", valor_servicos)
                            valor_iss = valores_nfse.get("ValorIss", valor_iss)

                        # Informações do tomador
                        tomador = None
                        if "DeclaracaoPrestacaoServico" in detalhes_nf:
                            tomador = detalhes_nf.get("DeclaracaoPrestacaoServico", {}).get("InfDeclaracaoPrestacaoServico", {}).get("Tomador", {})
                        else:
                            tomador = detalhes_nf.get("Tomador", {})

                        cliente = tomador.get("RazaoSocial", "Não informado") if tomador else "Não informado"

                        # CNPJ/CPF do tomador
                        cnpj_cpf_tomador = "Não informado"
                        if tomador and "IdentificacaoTomador" in tomador:
                            id_tomador = tomador["IdentificacaoTomador"]
                            if "CpfCnpj" in id_tomador:
                                cpf_cnpj = id_tomador["CpfCnpj"]
                                cnpj_cpf_tomador = cpf_cnpj.get("Cnpj", cpf_cnpj.get("Cpf", "Não informado"))

                        # Extrai campos adicionais
                        inscricao_municipal_prestador = "Não informado"
                        if "IdentificacaoPrestador" in prestador:
                            inscricao_municipal_prestador = prestador["IdentificacaoPrestador"].get("InscricaoMunicipal", "Não informado")

                        # Valores adicionais
                        base_calculo = "0.00"
                        if "ValoresNfse" in detalhes_nf:
                            valores_nfse = detalhes_nf["ValoresNfse"]
                            base_calculo = valores_nfse.get("BaseCalculo", valores_nfse.get("ValorServicos", "0.00"))
                            logging.debug(f"ValoresNfse encontrado: {valores_nfse}")

                        # Informações de ISS
                        iss_retido = "Não"
                        municipio_incidencia = "Não informado"
                        optante_simples = "Não"
                        incentivo_fiscal = "Não"

                        # Códigos de serviço
                        item_lista_servico = "Não informado"
                        codigo_cnae = "Não informado"
                        codigo_tributacao = "Não informado"

                        if "DeclaracaoPrestacaoServico" in detalhes_nf:
                            decl = detalhes_nf["DeclaracaoPrestacaoServico"]
                            if "InfDeclaracaoPrestacaoServico" in decl:
                                inf_decl = decl["InfDeclaracaoPrestacaoServico"]

                                # Informações fiscais
                                optante_simples = "Sim" if inf_decl.get("OptanteSimplesNacional", "0") == "1" else "Não"
                                incentivo_fiscal = "Sim" if inf_decl.get("IncentivoFiscal", "0") == "1" else "Não"

                                if "Servico" in inf_decl:
                                    servico = inf_decl["Servico"]

                                    # Códigos de serviço
                                    item_lista_servico = servico.get("ItemListaServico", "Não informado")
                                    codigo_cnae = servico.get("CodigoCnae", "Não informado")
                                    codigo_tributacao = servico.get("CodigoTributacaoMunicipio", "Não informado")

                                    # ISS
                                    iss_retido_codigo = servico.get("IssRetido", "0")
                                    iss_retido = "Sim" if iss_retido_codigo == "1" else "Não"

                                    # Município de incidência
                                    municipio_incidencia = servico.get("MunicipioIncidencia", "Não informado")

                        # Calcula o valor líquido (Base de Cálculo - Valor ISS)
                        try:
                            # Imprime os valores para debug
                            logging.debug(f"Base de Cálculo: {base_calculo}, tipo: {type(base_calculo)}")
                            logging.debug(f"Valor ISS: {valor_iss}, tipo: {type(valor_iss)}")

                            # Converte para float, tratando diferentes formatos
                            if isinstance(base_calculo, str):
                                base_calculo_num = float(base_calculo.replace(',', '.'))
                            else:
                                base_calculo_num = float(base_calculo) if base_calculo else 0.0

                            if isinstance(valor_iss, str):
                                valor_iss_num = float(valor_iss.replace(',', '.'))
                            else:
                                valor_iss_num = float(valor_iss) if valor_iss else 0.0

                            # Calcula o valor líquido
                            valor_liquido = base_calculo_num - valor_iss_num
                            valor_liquido = f"{valor_liquido:.2f}".replace('.', ',')

                            logging.debug(f"Valor Líquido calculado: {valor_liquido}")
                        except (ValueError, TypeError) as e:
                            logging.warning(f"Erro ao calcular valor líquido: {str(e)}")
                            logging.warning(f"Base de Cálculo: {base_calculo}, Valor ISS: {valor_iss}")
                            valor_liquido = "0,00"

                        # Adiciona todos os dados coletados
                        dados_coletados.append([
                            numero,
                            data_emissao,
                            competencia,
                            codigo_verificacao,
                            empresa,
                            cnpj_prestador,
                            inscricao_municipal_prestador,
                            cliente,
                            cnpj_cpf_tomador,
                            valor_servicos,
                            base_calculo,
                            aliquota,
                            valor_iss,
                            valor_liquido,
                            iss_retido,
                            item_lista_servico,
                            codigo_cnae,
                            codigo_tributacao,
                            discriminacao,
                            municipio,
                            uf,
                            municipio_incidencia,
                            optante_simples,
                            incentivo_fiscal
                        ])
                        return
                    except Exception as e:
                        logging.warning(f"Erro ao processar nota do arquivo {nota_xml}: {str(e)}")
                        return
            else:
                # Tenta encontrar a estrutura navegando pelo XML
                logging.warning(f"Estrutura desconhecida no arquivo {nota_xml}. Tentando identificar...")
                # Imprime as chaves de primeiro nível para debug
                logging.debug(f"Chaves de primeiro nível: {list(xml_dict.keys())}")

                # Se não conseguir identificar a estrutura, pula este arquivo
                logging.error(f"Não foi possível processar o arquivo {nota_xml}. Estrutura não reconhecida.")
                return

            # Extração de dados para NFe padrão
            numero = detalhes_nf.get("@Id", "Não informado")
            empresa = detalhes_nf.get('emit', {}).get('xNome', "Não informado")
            cliente = detalhes_nf.get("dest", {}).get("xNome", "Não informado")
            endereco_cliente = detalhes_nf.get("dest", {}).get("enderDest", "Endereço não informado")

            # Usa get() para evitar KeyError
            transp = detalhes_nf.get("transp", {})
            vol = transp.get("vol", {})
            # Se vol for uma lista, pega o primeiro item
            if isinstance(vol, list) and len(vol) > 0:
                peso_total = vol[0].get("pesoB", "Não informado")
            else:
                peso_total = vol.get("pesoB", "Não informado")

            dados_coletados.append([numero, empresa, cliente, endereco_cliente, peso_total])
            logging.info(f"Arquivo {nota_xml} processado com sucesso")

    except Exception as e:
        logging.error(f"Erro ao processar o arquivo {nota_xml}: {str(e)}")
        # Não interrompe o processamento, apenas registra o erro e continua

def main():
    logging.info("Iniciando processamento de notas fiscais XML")

    diretorio_saida = "Notas_Processadas"
    os.makedirs(diretorio_saida, exist_ok=True)

    try:
        arquivos_xml = os.listdir("nfs")
        if not arquivos_xml:
            logging.warning("Nenhum arquivo XML encontrado no diretório 'nfs'")
            print("Nenhum arquivo XML encontrado. Por favor, adicione arquivos XML ao diretório 'nfs'.")
            return
    except FileNotFoundError:
        logging.error("Diretório 'nfs' não encontrado")
        os.makedirs("nfs", exist_ok=True)
        print("Diretório 'nfs' não encontrado. Um novo diretório foi criado. Por favor, adicione arquivos XML a ele.")
        return

    # Colunas atualizadas com base na estrutura detalhada do XML de NFSe
    colunas_tabela = [
        "Número NF",
        "Data Emissão",
        "Competência",
        "Código Verificação",
        "Prestador",
        "CNPJ Prestador",
        "Inscrição Municipal Prestador",
        "Tomador",
        "CNPJ/CPF Tomador",
        "Valor Serviços",
        "Base de Cálculo",
        "Alíquota (%)",
        "Valor ISS",
        "Valor Líquido",
        "ISS Retido",
        "Item Lista Serviço",
        "Código CNAE",
        "Código Tributação Municipal",
        "Descrição Serviço",
        "Município Prestador",
        "UF Prestador",
        "Município Incidência",
        "Optante Simples Nacional",
        "Incentivo Fiscal"
    ]
    dados_extraidos = []

    arquivos_processados = 0
    arquivos_com_erro = 0

    for xml in arquivos_xml:
        if not xml.lower().endswith('.xml'):
            logging.warning(f"Arquivo {xml} não é um XML. Ignorando.")
            continue

        try:
            tamanho_antes = len(dados_extraidos)
            extrair_dados(xml, dados_extraidos)
            if len(dados_extraidos) > tamanho_antes:
                arquivos_processados += 1
            else:
                arquivos_com_erro += 1
        except Exception as e:
            arquivos_com_erro += 1
            logging.error(f"Erro não tratado ao processar {xml}: {str(e)}")

    if dados_extraidos:
        df_notas = pd.DataFrame(columns=colunas_tabela, data=dados_extraidos)

        # Verifica se o DataFrame é muito grande (mais de 100.000 linhas)
        tamanho_df = len(df_notas)
        if tamanho_df > 100000:
            logging.info(f"DataFrame muito grande ({tamanho_df} linhas). Processando em lotes para otimizar memória.")
            print(f"DataFrame muito grande ({tamanho_df} linhas). Processando em lotes para otimizar memória.")

            # Define o tamanho do lote
            tamanho_lote = 50000

            # Divide o DataFrame em lotes
            dfs_lotes = []
            for i in range(0, tamanho_df, tamanho_lote):
                fim = min(i + tamanho_lote, tamanho_df)
                dfs_lotes.append(df_notas.iloc[i:fim].copy())

            # Processa cada lote
            for i, df_lote in enumerate(dfs_lotes):
                logging.info(f"Processando lote {i+1} de {len(dfs_lotes)}...")

                # Formata as colunas de valores para exibir como moeda
                colunas_monetarias = ["Valor Serviços", "Base de Cálculo", "Alíquota (%)", "Valor ISS", "Valor Líquido"]
                for coluna in colunas_monetarias:
                    if coluna in df_lote.columns:
                        # Converte vírgula para ponto para garantir conversão correta para numérico
                        df_lote[coluna] = df_lote[coluna].astype(str).str.replace(',', '.').astype(float)

                        # Formata os valores monetários no padrão brasileiro (exceto Alíquota)
                        if coluna != "Alíquota (%)":
                            # Cria uma função para formatar os valores com pontos como separadores de milhar e vírgulas como separadores decimais
                            def formatar_moeda_br(x):
                                return f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

                            # Aplica a formatação
                            df_lote[coluna] = df_lote[coluna].apply(formatar_moeda_br)

                # Imprime algumas linhas para debug
                if i == 0:  # Apenas para o primeiro lote
                    logging.debug(f"Primeiras linhas do primeiro lote:\n{df_lote.head()}")

                # Formata a coluna de data - remove o fuso horário para compatibilidade com Excel
                if "Data Emissão" in df_lote.columns:
                    # Converte para datetime e remove o fuso horário
                    df_lote["Data Emissão"] = pd.to_datetime(df_lote["Data Emissão"], errors='coerce')
                    # Remove o fuso horário (timezone) das datas
                    df_lote["Data Emissão"] = df_lote["Data Emissão"].dt.tz_localize(None)

                # Formata a coluna de competência para MM/AAAA
                if "Competência" in df_lote.columns:
                    # Converte para datetime primeiro para garantir formato consistente
                    df_lote["Competência"] = pd.to_datetime(df_lote["Competência"], errors='coerce')
                    # Formata como MM/AAAA
                    df_lote["Competência"] = df_lote["Competência"].dt.strftime('%m/%Y')

                # Substitui o lote no DataFrame original
                df_notas.iloc[i*tamanho_lote:i*tamanho_lote+len(df_lote)] = df_lote

                # Libera memória
                del df_lote
                import gc
                gc.collect()
        else:
            # Formata as colunas de valores para exibir como moeda
            colunas_monetarias = ["Valor Serviços", "Base de Cálculo", "Alíquota (%)", "Valor ISS", "Valor Líquido"]
            for coluna in colunas_monetarias:
                if coluna in df_notas.columns:
                    # Converte vírgula para ponto para garantir conversão correta para numérico
                    df_notas[coluna] = df_notas[coluna].astype(str).str.replace(',', '.').astype(float)

                    # Formata os valores monetários no padrão brasileiro (exceto Alíquota)
                    if coluna != "Alíquota (%)":
                        # Cria uma função para formatar os valores com pontos como separadores de milhar e vírgulas como separadores decimais
                        def formatar_moeda_br(x):
                            return f"{x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

                        # Aplica a formatação
                        df_notas[coluna] = df_notas[coluna].apply(formatar_moeda_br)

            # Imprime algumas linhas para debug
            logging.debug(f"Primeiras linhas do DataFrame:\n{df_notas.head()}")

            # Formata a coluna de data - remove o fuso horário para compatibilidade com Excel
            if "Data Emissão" in df_notas.columns:
                # Converte para datetime e remove o fuso horário
                df_notas["Data Emissão"] = pd.to_datetime(df_notas["Data Emissão"], errors='coerce')
                # Remove o fuso horário (timezone) das datas
                df_notas["Data Emissão"] = df_notas["Data Emissão"].dt.tz_localize(None)

            # Formata a coluna de competência para MM/AAAA
            if "Competência" in df_notas.columns:
                # Converte para datetime primeiro para garantir formato consistente
                df_notas["Competência"] = pd.to_datetime(df_notas["Competência"], errors='coerce')
                # Formata como MM/AAAA
                df_notas["Competência"] = df_notas["Competência"].dt.strftime('%m/%Y')

        # Salva em Excel com formatação
        arquivo_saida = os.path.join(diretorio_saida, "NotasFiscais.xlsx")

        # Tenta salvar o arquivo Excel, com tratamento de erro para arquivo em uso
        try:
            # Verifica se o arquivo já existe e tenta criar um nome alternativo
            if os.path.exists(arquivo_saida):
                # Tenta fechar qualquer manipulador de arquivo que possa estar aberto
                import gc
                gc.collect()

                # Cria um nome de arquivo alternativo com timestamp
                timestamp = time.strftime("%Y%m%d_%H%M%S")
                arquivo_saida = os.path.join(diretorio_saida, f"NotasFiscais_{timestamp}.xlsx")
                logging.info(f"Arquivo existente, criando novo arquivo: {arquivo_saida}")

            # Verifica se o DataFrame é muito grande (mais de 100.000 linhas)
            if len(df_notas) > 100000:
                logging.info(f"Salvando DataFrame grande ({len(df_notas)} linhas) em Excel. Isso pode levar algum tempo...")
                print(f"Salvando DataFrame grande ({len(df_notas)} linhas) em Excel. Isso pode levar algum tempo...")

                # Usa o engine xlsxwriter para melhor performance com arquivos grandes
                df_notas.to_excel(arquivo_saida, index=False, sheet_name='Notas Fiscais', engine='xlsxwriter')
            else:
                # Para arquivos menores, usa o engine padrão
                df_notas.to_excel(arquivo_saida, index=False, sheet_name='Notas Fiscais')

            logging.info(f"Arquivo Excel salvo com sucesso: {arquivo_saida}")

            # Tenta aplicar formatação adicional apenas para arquivos menores
            if len(df_notas) <= 100000:
                try:
                    import openpyxl
                    from openpyxl.styles import Alignment, Font, PatternFill
                    from openpyxl.utils import get_column_letter

                    # Abre o arquivo Excel para formatação
                    wb = openpyxl.load_workbook(arquivo_saida)
                    ws = wb['Notas Fiscais']

                    # Insere uma linha no topo para os indicadores
                    ws.insert_rows(1)

                    # Calcula os indicadores
                    total_notas = len(df_notas)

                    # Precisamos converter as colunas monetárias de volta para float para calcular os totais
                    colunas_monetarias = ["Base de Cálculo", "Valor ISS", "Valor Líquido"]
                    df_temp = df_notas.copy()

                    for coluna in colunas_monetarias:
                        if coluna in df_temp.columns:
                            # Verifica se a coluna contém strings (já formatadas)
                            if df_temp[coluna].dtype == 'object':
                                # Converte de volta para float para calcular a soma
                                df_temp[coluna] = df_temp[coluna].str.replace('.', '').str.replace(',', '.').astype(float)

                    # Calcula os totais das colunas monetárias
                    total_base_calculo = df_temp["Base de Cálculo"].sum() if "Base de Cálculo" in df_temp.columns else 0
                    total_valor_iss = df_temp["Valor ISS"].sum() if "Valor ISS" in df_temp.columns else 0
                    total_valor_liquido = df_temp["Valor Líquido"].sum() if "Valor Líquido" in df_temp.columns else 0

                    # Função para formatar valores monetários no padrão brasileiro
                    def formatar_valor_br(valor):
                        # Formata com ponto como separador de milhar e vírgula como separador decimal
                        return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

                    # Adiciona os indicadores na primeira linha
                    ws.cell(row=1, column=1).value = f"Total de Notas: {total_notas}"

                    # Encontra os índices das colunas monetárias
                    base_calculo_idx = None
                    valor_iss_idx = None
                    valor_liquido_idx = None

                    for idx, col in enumerate(df_notas.columns):
                        if col == "Base de Cálculo":
                            base_calculo_idx = idx + 1  # +1 porque as colunas do Excel começam em 1
                            ws.cell(row=1, column=base_calculo_idx).value = f"Total Base de Cálculo: {formatar_valor_br(total_base_calculo)}"
                        elif col == "Valor ISS":
                            valor_iss_idx = idx + 1
                            ws.cell(row=1, column=valor_iss_idx).value = f"Total Valor ISS: {formatar_valor_br(total_valor_iss)}"
                        elif col == "Valor Líquido":
                            valor_liquido_idx = idx + 1
                            ws.cell(row=1, column=valor_liquido_idx).value = f"Total Valor Líquido: {formatar_valor_br(total_valor_liquido)}"

                    # Formata a linha de indicadores
                    for col in range(1, len(df_notas.columns) + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")  # Cinza claro

                    # Formata o cabeçalho (agora na linha 2)
                    for col in range(1, len(df_notas.columns) + 1):
                        cell = ws.cell(row=2, column=col)
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    # Ajusta a largura das colunas
                    for idx, col in enumerate(df_notas.columns):
                        # Encontra o comprimento máximo na coluna (amostra para performance)
                        if len(df_notas) > 1000:
                            # Para DataFrames grandes, usa apenas uma amostra para calcular a largura
                            amostra = df_notas[col].sample(n=min(1000, len(df_notas)))
                            max_len = max(
                                amostra.astype(str).apply(len).max(),
                                len(str(col))
                            ) + 2
                        else:
                            max_len = max(
                                df_notas[col].astype(str).apply(len).max(),
                                len(str(col))
                            ) + 2

                        # Define a largura da coluna
                        column_letter = get_column_letter(idx + 1)
                        ws.column_dimensions[column_letter].width = min(max_len, 50)  # Limita a 50 caracteres

                    # Adiciona uma nova aba "Indicadores"
                    if "Indicadores" in wb.sheetnames:
                        ws_indicadores = wb["Indicadores"]
                    else:
                        ws_indicadores = wb.create_sheet("Indicadores")

                    # Função para adicionar título de seção
                    def adicionar_titulo_secao(ws, row, texto):
                        ws.cell(row=row, column=1).value = texto
                        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                        ws.merge_cells(f'A{row}:C{row}')
                        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')
                        ws.cell(row=row, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                        return row + 1

                    # Função para adicionar cabeçalhos
                    def adicionar_cabecalhos(ws, row):
                        ws.cell(row=row, column=1).value = "Indicador"
                        ws.cell(row=row, column=2).value = "Valor"
                        ws.cell(row=row, column=3).value = "Observação"

                        for col in range(1, 4):
                            ws.cell(row=row, column=col).font = Font(bold=True)
                            ws.cell(row=row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
                        return row + 1

                    # Função para adicionar indicador
                    def adicionar_indicador(ws, row, indicador, valor, obs, i):
                        ws.cell(row=row, column=1).value = indicador
                        ws.cell(row=row, column=2).value = valor
                        ws.cell(row=row, column=3).value = obs

                        # Formata células
                        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left')

                        # Verifica o tipo de valor para definir o alinhamento apropriado
                        if isinstance(valor, (int, float)) or (isinstance(valor, str) and (valor.startswith('R$') or '%' in valor or ',' in valor)):
                            # Valores numéricos, monetários ou percentuais são alinhados à direita
                            ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')
                        else:
                            # Valores de texto (como nomes de municípios, UFs) são alinhados à esquerda
                            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')

                        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')

                        # Adiciona cor de fundo alternada para facilitar a leitura
                        if i % 2 == 0:
                            for col in range(1, 4):
                                ws.cell(row=row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                        # Adiciona borda
                        for col in range(1, 4):
                            ws.cell(row=row, column=col).border = thin_border

                        return row + 1

                    # Adiciona título principal na aba de indicadores
                    ws_indicadores.cell(row=1, column=1).value = "PAINEL DE INDICADORES"
                    ws_indicadores.cell(row=1, column=1).font = Font(bold=True, size=14)
                    ws_indicadores.merge_cells('A1:C1')
                    ws_indicadores.cell(row=1, column=1).alignment = Alignment(horizontal='center')

                    # Prepara borda para todas as células
                    from openpyxl.styles import Border, Side
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    # Ajusta largura das colunas
                    ws_indicadores.column_dimensions['A'].width = 30
                    ws_indicadores.column_dimensions['B'].width = 30
                    ws_indicadores.column_dimensions['C'].width = 45

                    # Prepara DataFrame para cálculos
                    # Converte colunas de data para datetime
                    if "Data Emissão" in df_notas.columns:
                        df_notas["Data Emissão"] = pd.to_datetime(df_notas["Data Emissão"], errors='coerce')

                    # Calcula indicadores básicos
                    media_base_calculo = total_base_calculo / total_notas if total_notas > 0 else 0
                    media_valor_iss = total_valor_iss / total_notas if total_notas > 0 else 0
                    media_valor_liquido = total_valor_liquido / total_notas if total_notas > 0 else 0

                    # Calcula valores máximos e mínimos
                    max_base_calculo = df_temp["Base de Cálculo"].max() if "Base de Cálculo" in df_temp.columns else 0
                    min_base_calculo = df_temp["Base de Cálculo"].min() if "Base de Cálculo" in df_temp.columns else 0
                    max_valor_iss = df_temp["Valor ISS"].max() if "Valor ISS" in df_temp.columns else 0
                    min_valor_iss = df_temp["Valor ISS"].min() if "Valor ISS" in df_temp.columns else 0
                    max_valor_liquido = df_temp["Valor Líquido"].max() if "Valor Líquido" in df_temp.columns else 0
                    min_valor_liquido = df_temp["Valor Líquido"].min() if "Valor Líquido" in df_temp.columns else 0

                    # Calcula desvio padrão e mediana
                    desvio_padrao_base_calculo = df_temp["Base de Cálculo"].std() if "Base de Cálculo" in df_temp.columns else 0
                    mediana_base_calculo = df_temp["Base de Cálculo"].median() if "Base de Cálculo" in df_temp.columns else 0

                    # Calcula indicadores temporais
                    primeira_nota_data = df_notas["Data Emissão"].min() if "Data Emissão" in df_notas.columns else None
                    ultima_nota_data = df_notas["Data Emissão"].max() if "Data Emissão" in df_notas.columns else None

                    # Calcula período em dias
                    periodo_dias = (ultima_nota_data - primeira_nota_data).days + 1 if primeira_nota_data and ultima_nota_data else 0
                    media_notas_por_dia = total_notas / periodo_dias if periodo_dias > 0 else 0

                    # Calcula indicadores por tomador de serviço
                    total_clientes_unicos = df_notas["Tomador"].nunique() if "Tomador" in df_notas.columns else 0

                    # Tomador com maior volume financeiro
                    if "Tomador" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        # Cria um DataFrame temporário com tomador e base de cálculo
                        df_cliente_valor = pd.DataFrame({
                            'Cliente': df_notas['Tomador'],
                            'Base de Cálculo': df_temp['Base de Cálculo']
                        })

                        # Agrupa por tomador e soma os valores
                        clientes_por_valor = df_cliente_valor.groupby('Cliente')['Base de Cálculo'].sum().sort_values(ascending=False)

                        if not clientes_por_valor.empty:
                            cliente_maior_valor = clientes_por_valor.index[0]
                            valor_maior_cliente = clientes_por_valor.iloc[0]

                            # Top 5 clientes por valor
                            top5_clientes = clientes_por_valor.head(5)
                            top5_clientes_nomes = list(top5_clientes.index)
                            top5_clientes_valores = list(top5_clientes.values)

                            # Calcula quantos clientes representam 80% do faturamento
                            faturamento_acumulado = 0
                            clientes_80_porcento = 0
                            for valor in clientes_por_valor:
                                faturamento_acumulado += valor
                                clientes_80_porcento += 1
                                if faturamento_acumulado >= (total_base_calculo * 0.8):
                                    break
                        else:
                            cliente_maior_valor = "N/A"
                            valor_maior_cliente = 0
                            top5_clientes_nomes = ["N/A"] * 5
                            top5_clientes_valores = [0] * 5
                            clientes_80_porcento = 0

                        # Tomador com maior quantidade de notas
                        clientes_por_qtd = df_notas.groupby('Tomador').size().sort_values(ascending=False)

                        if not clientes_por_qtd.empty:
                            cliente_maior_qtd = clientes_por_qtd.index[0]
                            qtd_maior_cliente = clientes_por_qtd.iloc[0]

                            # Top 5 clientes por quantidade
                            top5_clientes_qtd = clientes_por_qtd.head(5)
                            top5_clientes_qtd_nomes = list(top5_clientes_qtd.index)
                            top5_clientes_qtd_valores = list(top5_clientes_qtd.values)
                        else:
                            cliente_maior_qtd = "N/A"
                            qtd_maior_cliente = 0
                            top5_clientes_qtd_nomes = ["N/A"] * 5
                            top5_clientes_qtd_valores = [0] * 5

                        # Concentração de faturamento (top 5 clientes)
                        if not clientes_por_valor.empty:
                            top5_clientes_valor = clientes_por_valor.head(5).sum()
                            concentracao_top5 = (top5_clientes_valor / total_base_calculo) * 100 if total_base_calculo > 0 else 0
                        else:
                            concentracao_top5 = 0
                    else:
                        cliente_maior_valor = "N/A"
                        valor_maior_cliente = 0
                        top5_clientes_nomes = ["N/A"] * 5
                        top5_clientes_valores = [0] * 5
                        cliente_maior_qtd = "N/A"
                        qtd_maior_cliente = 0
                        top5_clientes_qtd_nomes = ["N/A"] * 5
                        top5_clientes_qtd_valores = [0] * 5
                        concentracao_top5 = 0
                        clientes_80_porcento = 0

                    # Calcula indicadores fiscais e tributários
                    if "Alíquota (%)" in df_temp.columns:
                        aliquota_media = df_temp["Alíquota (%)"].mean()
                    else:
                        aliquota_media = 0

                    # Percentual de notas com ISS retido
                    if "ISS Retido" in df_notas.columns:
                        notas_iss_retido = df_notas[df_notas["ISS Retido"] == "Sim"].shape[0]
                        percentual_iss_retido = (notas_iss_retido / total_notas) * 100 if total_notas > 0 else 0
                    else:
                        notas_iss_retido = 0
                        percentual_iss_retido = 0

                    # Percentual de notas para optantes do Simples Nacional
                    if "Optante Simples Nacional" in df_notas.columns:
                        notas_simples = df_notas[df_notas["Optante Simples Nacional"] == "Sim"].shape[0]
                        percentual_simples = (notas_simples / total_notas) * 100 if total_notas > 0 else 0
                    else:
                        notas_simples = 0
                        percentual_simples = 0

                    # Calcula indicadores de serviços
                    if "Item Lista Serviço" in df_notas.columns:
                        # Serviços mais prestados
                        servicos_mais_prestados = df_notas.groupby("Item Lista Serviço").size().sort_values(ascending=False)

                        if not servicos_mais_prestados.empty:
                            servico_mais_comum = servicos_mais_prestados.index[0]
                            qtd_servico_mais_comum = servicos_mais_prestados.iloc[0]
                        else:
                            servico_mais_comum = "N/A"
                            qtd_servico_mais_comum = 0
                    else:
                        servico_mais_comum = "N/A"
                        qtd_servico_mais_comum = 0

                    # Calcula indicadores geográficos
                    if "Município" in df_notas.columns:
                        # Distribuição por município
                        municipios = df_notas.groupby("Município").size().sort_values(ascending=False)

                        if not municipios.empty:
                            municipio_mais_comum = municipios.index[0]
                            qtd_municipio_mais_comum = municipios.iloc[0]

                            # Top 5 municípios por quantidade
                            top5_municipios = municipios.head(5)
                            top5_municipios_nomes = list(top5_municipios.index)
                            top5_municipios_valores = list(top5_municipios.values)

                            # Calcula valor por município
                            if "Base de Cálculo" in df_temp.columns:
                                df_mun_valor = pd.DataFrame({
                                    'Município': df_notas['Município'],
                                    'Base de Cálculo': df_temp['Base de Cálculo']
                                })

                                # Agrupa por município e soma os valores
                                municipios_por_valor = df_mun_valor.groupby('Município')['Base de Cálculo'].sum().sort_values(ascending=False)

                                if not municipios_por_valor.empty:
                                    municipio_maior_valor = municipios_por_valor.index[0]
                                    valor_municipio_maior = municipios_por_valor.iloc[0]

                                    # Top 5 municípios por valor
                                    top5_municipios_valor = municipios_por_valor.head(5)
                                    top5_municipios_valor_nomes = list(top5_municipios_valor.index)
                                    top5_municipios_valor_valores = list(top5_municipios_valor.values)

                                    # Concentração por município (percentual do top 5)
                                    concentracao_top5_municipios = (top5_municipios_valor.sum() / total_base_calculo) * 100 if total_base_calculo > 0 else 0
                                else:
                                    municipio_maior_valor = "N/A"
                                    valor_municipio_maior = 0
                                    top5_municipios_valor_nomes = ["N/A"] * 5
                                    top5_municipios_valor_valores = [0] * 5
                                    concentracao_top5_municipios = 0
                            else:
                                municipio_maior_valor = "N/A"
                                valor_municipio_maior = 0
                                top5_municipios_valor_nomes = ["N/A"] * 5
                                top5_municipios_valor_valores = [0] * 5
                                concentracao_top5_municipios = 0
                        else:
                            municipio_mais_comum = "N/A"
                            qtd_municipio_mais_comum = 0
                            top5_municipios_nomes = ["N/A"] * 5
                            top5_municipios_valores = [0] * 5
                            municipio_maior_valor = "N/A"
                            valor_municipio_maior = 0
                            top5_municipios_valor_nomes = ["N/A"] * 5
                            top5_municipios_valor_valores = [0] * 5
                            concentracao_top5_municipios = 0
                    else:
                        municipio_mais_comum = "N/A"
                        qtd_municipio_mais_comum = 0
                        top5_municipios_nomes = ["N/A"] * 5
                        top5_municipios_valores = [0] * 5
                        municipio_maior_valor = "N/A"
                        valor_municipio_maior = 0
                        top5_municipios_valor_nomes = ["N/A"] * 5
                        top5_municipios_valor_valores = [0] * 5
                        concentracao_top5_municipios = 0

                    if "UF" in df_notas.columns:
                        # Distribuição por UF
                        ufs = df_notas.groupby("UF").size().sort_values(ascending=False)

                        if not ufs.empty:
                            uf_mais_comum = ufs.index[0]
                            qtd_uf_mais_comum = ufs.iloc[0]

                            # Top 5 UFs por quantidade
                            top5_ufs = ufs.head(min(5, len(ufs)))
                            top5_ufs_nomes = list(top5_ufs.index)
                            top5_ufs_valores = list(top5_ufs.values)

                            # Calcula valor por UF
                            if "Base de Cálculo" in df_temp.columns:
                                df_uf_valor = pd.DataFrame({
                                    'UF': df_notas['UF'],
                                    'Base de Cálculo': df_temp['Base de Cálculo']
                                })

                                # Agrupa por UF e soma os valores
                                ufs_por_valor = df_uf_valor.groupby('UF')['Base de Cálculo'].sum().sort_values(ascending=False)

                                if not ufs_por_valor.empty:
                                    uf_maior_valor = ufs_por_valor.index[0]
                                    valor_uf_maior = ufs_por_valor.iloc[0]

                                    # Top 5 UFs por valor (ou menos se houver menos de 5 UFs)
                                    top5_ufs_valor = ufs_por_valor.head(min(5, len(ufs_por_valor)))
                                    top5_ufs_valor_nomes = list(top5_ufs_valor.index)
                                    top5_ufs_valor_valores = list(top5_ufs_valor.values)

                                    # Valor médio por UF
                                    valor_medio_por_uf = total_base_calculo / len(ufs_por_valor) if len(ufs_por_valor) > 0 else 0
                                else:
                                    uf_maior_valor = "N/A"
                                    valor_uf_maior = 0
                                    top5_ufs_valor_nomes = ["N/A"] * min(5, len(ufs))
                                    top5_ufs_valor_valores = [0] * min(5, len(ufs))
                                    valor_medio_por_uf = 0
                            else:
                                uf_maior_valor = "N/A"
                                valor_uf_maior = 0
                                top5_ufs_valor_nomes = ["N/A"] * min(5, len(ufs))
                                top5_ufs_valor_valores = [0] * min(5, len(ufs))
                                valor_medio_por_uf = 0
                        else:
                            uf_mais_comum = "N/A"
                            qtd_uf_mais_comum = 0
                            top5_ufs_nomes = ["N/A"] * 5
                            top5_ufs_valores = [0] * 5
                            uf_maior_valor = "N/A"
                            valor_uf_maior = 0
                            top5_ufs_valor_nomes = ["N/A"] * 5
                            top5_ufs_valor_valores = [0] * 5
                            valor_medio_por_uf = 0
                    else:
                        uf_mais_comum = "N/A"
                        qtd_uf_mais_comum = 0
                        top5_ufs_nomes = ["N/A"] * 5
                        top5_ufs_valores = [0] * 5
                        uf_maior_valor = "N/A"
                        valor_uf_maior = 0
                        top5_ufs_valor_nomes = ["N/A"] * 5
                        top5_ufs_valor_valores = [0] * 5
                        valor_medio_por_uf = 0

                    # Adiciona os indicadores na planilha por seções
                    current_row = 3

                    # Seção 1: Indicadores Gerais
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "1. INDICADORES GERAIS")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    indicadores_gerais = [
                        ("Total de Notas Fiscais", total_notas, "Quantidade total de notas processadas"),
                        ("Total Base de Cálculo", formatar_valor_br(total_base_calculo), "Soma de todas as bases de cálculo"),
                        ("Total Valor ISS", formatar_valor_br(total_valor_iss), "Soma de todos os valores de ISS"),
                        ("Total Valor Líquido", formatar_valor_br(total_valor_liquido), "Soma de todos os valores líquidos"),
                        ("Média Base de Cálculo", formatar_valor_br(media_base_calculo), "Valor médio da base de cálculo por nota"),
                        ("Média Valor ISS", formatar_valor_br(media_valor_iss), "Valor médio do ISS por nota"),
                        ("Média Valor Líquido", formatar_valor_br(media_valor_liquido), "Valor médio líquido por nota"),
                        ("Desvio Padrão Base de Cálculo", formatar_valor_br(desvio_padrao_base_calculo), "Medida de dispersão dos valores"),
                        ("Mediana Base de Cálculo", formatar_valor_br(mediana_base_calculo), "Valor central (50% acima, 50% abaixo)"),
                        ("Máximo Base de Cálculo", formatar_valor_br(max_base_calculo), "Maior valor de base de cálculo"),
                        ("Mínimo Base de Cálculo", formatar_valor_br(min_base_calculo), "Menor valor de base de cálculo"),
                        ("Máximo Valor ISS", formatar_valor_br(max_valor_iss), "Maior valor de ISS"),
                        ("Mínimo Valor ISS", formatar_valor_br(min_valor_iss), "Menor valor de ISS"),
                        ("Máximo Valor Líquido", formatar_valor_br(max_valor_liquido), "Maior valor líquido"),
                        ("Mínimo Valor Líquido", formatar_valor_br(min_valor_liquido), "Menor valor líquido"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_gerais):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # Adiciona espaço entre seções
                    current_row += 1

                    # Seção 2: Indicadores Temporais
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "2. INDICADORES TEMPORAIS")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Formata datas para exibição
                    primeira_nota_str = primeira_nota_data.strftime('%d/%m/%Y') if primeira_nota_data else "N/A"
                    ultima_nota_str = ultima_nota_data.strftime('%d/%m/%Y') if ultima_nota_data else "N/A"

                    indicadores_temporais = [
                        ("Período de Análise", f"{primeira_nota_str} a {ultima_nota_str}", "Intervalo entre a primeira e última nota"),
                        ("Duração do Período", f"{periodo_dias} dias", "Número de dias entre a primeira e última nota"),
                        ("Média de Notas por Dia", f"{media_notas_por_dia:.2f}", "Média diária de notas emitidas no período"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_temporais):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # Adiciona espaço entre seções
                    current_row += 1

                    # Seção 3: Análise de Tomadores e Prestadores
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "3. ANÁLISE DE TOMADORES E PRESTADORES")

                    # 3.1 Indicadores Gerais de Tomadores
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "3.1 Indicadores Gerais de Tomadores")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula a quantidade de prestadores únicos
                    total_prestadores_unicos = df_notas["Prestador"].nunique() if "Prestador" in df_notas.columns else 0

                    # Calcula o valor médio por tomador
                    valor_medio_por_tomador = total_base_calculo / total_clientes_unicos if total_clientes_unicos > 0 else 0

                    # Calcula a quantidade média de notas por tomador
                    notas_por_tomador = total_notas / total_clientes_unicos if total_clientes_unicos > 0 else 0

                    # Calcula a relação tomador/prestador
                    relacao_tomador_prestador = total_clientes_unicos / total_prestadores_unicos if total_prestadores_unicos > 0 else 0

                    indicadores_gerais_tomador = [
                        ("Quantidade de Tomadores Únicos", total_clientes_unicos, "Número total de tomadores de serviço diferentes"),
                        ("Valor Médio por Tomador", formatar_valor_br(valor_medio_por_tomador), "Valor médio dos serviços prestados a cada tomador"),
                        ("Quantidade Média de Notas por Tomador", f"{notas_por_tomador:.2f}", "Média de notas fiscais emitidas para cada tomador"),
                        ("Relação Tomador/Prestador", f"{relacao_tomador_prestador:.2f}", "Quantidade de tomadores para cada prestador"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_gerais_tomador):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # 3.2 Indicadores de Concentração
                    current_row += 1
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "3.2 Indicadores de Concentração")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula o índice de Gini (aproximação simples)
                    if not clientes_por_valor.empty and len(clientes_por_valor) > 1:
                        # Normaliza os valores
                        valores_normalizados = clientes_por_valor.values / clientes_por_valor.sum()
                        # Ordena os valores (sem usar numpy)
                        valores_ordenados = sorted(valores_normalizados)
                        # Calcula o índice de Gini (sem usar numpy)
                        n = len(valores_ordenados)
                        soma_ponderada = sum([(n + 1 - i) * valores_ordenados[i-1] for i in range(1, n + 1)])
                        indice_gini = 1 - (2 * soma_ponderada) / (n * sum(valores_ordenados))
                    else:
                        indice_gini = 0

                    # Calcula o percentual do maior tomador
                    percentual_maior_tomador = (valor_maior_cliente / total_base_calculo) * 100 if total_base_calculo > 0 else 0

                    # Calcula a quantidade de notas emitidas para o tomador com maior volume
                    qtd_notas_maior_tomador = 0
                    if "Tomador" in df_notas.columns and cliente_maior_valor != "N/A":
                        qtd_notas_maior_tomador = df_notas[df_notas["Tomador"] == cliente_maior_valor].shape[0]

                    indicadores_concentracao = [
                        ("Tomador com Maior Volume", cliente_maior_valor, "Tomador para o qual foi emitido maior valor em notas"),
                        ("Valor do Tomador com Maior Volume", formatar_valor_br(valor_maior_cliente), "Valor total dos serviços prestados ao tomador de maior volume"),
                        ("Quantidade de Notas do Tomador com Maior Volume", qtd_notas_maior_tomador, "Número de notas emitidas para o tomador com maior volume"),
                        ("Percentual do Maior Tomador", f"{percentual_maior_tomador:.2f}%", "Percentual do faturamento representado pelo maior tomador"),
                        ("Concentração Top 5 Tomadores", f"{concentracao_top5:.2f}%", "Percentual do faturamento dos 5 maiores tomadores"),
                        ("Tomadores para 80% do Faturamento", clientes_80_porcento, "Quantidade de tomadores que representam 80% do faturamento"),
                        ("Índice de Concentração (Gini)", f"{indice_gini:.4f}", "Medida de desigualdade (0=igualdade perfeita, 1=concentração total)"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_concentracao):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # 3.3 Indicadores de Frequência
                    current_row += 1
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "3.3 Indicadores de Frequência")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula tomadores por frequência de notas
                    if "Tomador" in df_notas.columns:
                        # Contagem de notas por tomador
                        notas_por_tomador_df = df_notas.groupby('Tomador').size().reset_index(name='Quantidade')

                        # Tomadores com apenas 1 nota
                        tomadores_1_nota = notas_por_tomador_df[notas_por_tomador_df['Quantidade'] == 1].shape[0]
                        percentual_tomadores_1_nota = (tomadores_1_nota / total_clientes_unicos) * 100 if total_clientes_unicos > 0 else 0

                        # Tomadores com 2 a 5 notas
                        tomadores_2_5_notas = notas_por_tomador_df[(notas_por_tomador_df['Quantidade'] >= 2) & (notas_por_tomador_df['Quantidade'] <= 5)].shape[0]
                        percentual_tomadores_2_5_notas = (tomadores_2_5_notas / total_clientes_unicos) * 100 if total_clientes_unicos > 0 else 0

                        # Tomadores com mais de 5 notas
                        tomadores_mais_5_notas = notas_por_tomador_df[notas_por_tomador_df['Quantidade'] > 5].shape[0]
                        percentual_tomadores_mais_5_notas = (tomadores_mais_5_notas / total_clientes_unicos) * 100 if total_clientes_unicos > 0 else 0
                    else:
                        tomadores_1_nota = 0
                        percentual_tomadores_1_nota = 0
                        tomadores_2_5_notas = 0
                        percentual_tomadores_2_5_notas = 0
                        tomadores_mais_5_notas = 0
                        percentual_tomadores_mais_5_notas = 0

                    indicadores_frequencia = [
                        ("Tomador com Maior Quantidade", cliente_maior_qtd, "Tomador para o qual foram emitidas mais notas fiscais"),
                        ("Quantidade de Notas do Tomador", qtd_maior_cliente, "Número de notas emitidas para o tomador com maior quantidade"),
                        ("Tomadores com Apenas 1 Nota", f"{tomadores_1_nota} ({percentual_tomadores_1_nota:.2f}%)", "Quantidade e percentual de tomadores para os quais foi emitida apenas 1 nota"),
                        ("Tomadores com 2 a 5 Notas", f"{tomadores_2_5_notas} ({percentual_tomadores_2_5_notas:.2f}%)", "Quantidade e percentual de tomadores para os quais foram emitidas de 2 a 5 notas"),
                        ("Tomadores com Mais de 5 Notas", f"{tomadores_mais_5_notas} ({percentual_tomadores_mais_5_notas:.2f}%)", "Quantidade e percentual de tomadores para os quais foram emitidas mais de 5 notas"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_frequencia):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # 3.4 Indicadores de Valor por Nota
                    current_row += 1
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "3.4 Indicadores de Valor por Nota")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula o valor médio por nota para cada tomador
                    if "Tomador" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        # Cria um DataFrame com tomador e base de cálculo
                        df_tomador_valor = pd.DataFrame({
                            'Tomador': df_notas['Tomador'],
                            'Base de Cálculo': df_temp['Base de Cálculo']
                        })

                        # Calcula a soma e contagem por tomador
                        valor_por_tomador = df_tomador_valor.groupby('Tomador')['Base de Cálculo'].sum()
                        notas_por_tomador_count = df_tomador_valor.groupby('Tomador').size()

                        # Calcula o valor médio por nota para cada tomador
                        valor_medio_por_nota = valor_por_tomador / notas_por_tomador_count

                        # Encontra o tomador com maior valor médio por nota
                        if not valor_medio_por_nota.empty:
                            tomador_maior_valor_medio = valor_medio_por_nota.idxmax()
                            valor_medio_maior = valor_medio_por_nota.max()

                            # Encontra o tomador com menor valor médio por nota
                            tomador_menor_valor_medio = valor_medio_por_nota.idxmin()
                            valor_medio_menor = valor_medio_por_nota.min()

                            # Calcula o desvio padrão dos valores médios
                            desvio_padrao_valores_medios = valor_medio_por_nota.std()
                        else:
                            tomador_maior_valor_medio = "N/A"
                            valor_medio_maior = 0
                            tomador_menor_valor_medio = "N/A"
                            valor_medio_menor = 0
                            desvio_padrao_valores_medios = 0
                    else:
                        tomador_maior_valor_medio = "N/A"
                        valor_medio_maior = 0
                        tomador_menor_valor_medio = "N/A"
                        valor_medio_menor = 0
                        desvio_padrao_valores_medios = 0

                    indicadores_valor_nota = [
                        ("Tomador com Maior Valor Médio por Nota", tomador_maior_valor_medio, "Tomador com maior valor médio por nota fiscal"),
                        ("Valor Médio por Nota", formatar_valor_br(valor_medio_maior), "Valor médio por nota do tomador com maior média"),
                        ("Tomador com Menor Valor Médio por Nota", tomador_menor_valor_medio, "Tomador com menor valor médio por nota fiscal"),
                        ("Valor Médio por Nota", formatar_valor_br(valor_medio_menor), "Valor médio por nota do tomador com menor média"),
                        ("Desvio Padrão dos Valores Médios", formatar_valor_br(desvio_padrao_valores_medios), "Variação dos valores médios entre tomadores"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_valor_nota):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # 3.5 Tabelas Detalhadas de Tomadores
                    current_row += 1
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "3.5 Tabelas Detalhadas de Tomadores")

                    # Adiciona os top 5 tomadores por valor
                    if "Tomador" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        # Adiciona subtítulo para os top 5 tomadores
                        current_row += 1
                        ws_indicadores.cell(row=current_row, column=1).value = "Top 5 Tomadores por Valor"
                        ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                        ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                        ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                        current_row += 1

                        # Adiciona cabeçalhos para os top 5 tomadores
                        ws_indicadores.cell(row=current_row, column=1).value = "Tomador"
                        ws_indicadores.cell(row=current_row, column=2).value = "Valor Total"
                        ws_indicadores.cell(row=current_row, column=3).value = "% do Faturamento"

                        for col in range(1, 4):
                            ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                            ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                            ws_indicadores.cell(row=current_row, column=col).border = thin_border
                        current_row += 1

                        # Adiciona os top 5 tomadores
                        tomadores_por_valor = df_tomador_valor.groupby('Tomador')['Base de Cálculo'].sum().sort_values(ascending=False)
                        top5_tomadores = tomadores_por_valor.head(5)

                        for i, (tomador, valor) in enumerate(top5_tomadores.items()):
                            percentual = (valor / total_base_calculo) * 100 if total_base_calculo > 0 else 0

                            ws_indicadores.cell(row=current_row, column=1).value = tomador
                            ws_indicadores.cell(row=current_row, column=2).value = formatar_valor_br(valor)
                            ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                            # Formata células
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                            ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                            # Adiciona cor de fundo alternada para facilitar a leitura
                            if i % 2 == 0:
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                            # Adiciona borda
                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border

                            current_row += 1

                    # Adiciona espaço entre seções
                    current_row += 1

                    # Seção 4: Indicadores Fiscais e Tributários
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "4. INDICADORES FISCAIS E TRIBUTÁRIOS")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    indicadores_fiscais = [
                        ("Alíquota Média de ISS", f"{aliquota_media:.2f}%", "Média das alíquotas aplicadas"),
                        ("Notas com ISS Retido", f"{notas_iss_retido} ({percentual_iss_retido:.2f}%)", "Quantidade e percentual de notas com ISS retido"),
                        ("Notas para Optantes do Simples", f"{notas_simples} ({percentual_simples:.2f}%)", "Quantidade e percentual de notas para optantes do Simples"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_fiscais):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # Adiciona espaço entre seções
                    current_row += 1

                    # Seção 5: Indicadores de Serviços
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "5. INDICADORES DE SERVIÇOS")

                    # 5.1 Indicadores Gerais de Serviços
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "5.1 Indicadores Gerais de Serviços")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula a quantidade de serviços únicos
                    total_servicos_unicos = df_notas["Item Lista Serviço"].nunique() if "Item Lista Serviço" in df_notas.columns else 0

                    # Calcula o percentual do serviço mais comum
                    percentual_servico_mais_comum = (qtd_servico_mais_comum / total_notas) * 100 if total_notas > 0 else 0

                    indicadores_gerais_servicos = [
                        ("Serviço Mais Prestado", f"{servico_mais_comum} ({qtd_servico_mais_comum} notas)", "Código de serviço mais frequente"),
                        ("Percentual do Serviço Mais Prestado", f"{percentual_servico_mais_comum:.2f}%", "Percentual do serviço mais frequente em relação ao total de notas"),
                        ("Quantidade de Serviços Únicos", total_servicos_unicos, "Número total de códigos de serviço diferentes"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_gerais_servicos):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # 5.2 Análise de Valor por Serviço
                    current_row += 1
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "5.2 Análise de Valor por Serviço")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula o valor por serviço
                    if "Item Lista Serviço" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        # Cria um DataFrame com serviço e base de cálculo
                        df_servico_valor = pd.DataFrame({
                            'Serviço': df_notas['Item Lista Serviço'],
                            'Base de Cálculo': df_temp['Base de Cálculo']
                        })

                        # Agrupa por serviço e soma os valores
                        servicos_por_valor = df_servico_valor.groupby('Serviço')['Base de Cálculo'].sum().sort_values(ascending=False)

                        if not servicos_por_valor.empty:
                            # Serviço com maior valor total
                            servico_maior_valor = servicos_por_valor.index[0]
                            valor_maior_servico = servicos_por_valor.iloc[0]

                            # Percentual do serviço com maior valor
                            percentual_maior_servico = (valor_maior_servico / total_base_calculo) * 100 if total_base_calculo > 0 else 0

                            # Calcula a quantidade de notas do serviço com maior valor
                            qtd_notas_maior_servico = df_notas[df_notas["Item Lista Serviço"] == servico_maior_valor].shape[0]

                            # Calcula o valor médio por nota para o serviço com maior valor
                            valor_medio_por_nota_maior_servico = valor_maior_servico / qtd_notas_maior_servico if qtd_notas_maior_servico > 0 else 0

                            # Calcula a concentração dos top 5 serviços
                            top5_servicos_valor = servicos_por_valor.head(5).sum()
                            concentracao_top5_servicos = (top5_servicos_valor / total_base_calculo) * 100 if total_base_calculo > 0 else 0
                        else:
                            servico_maior_valor = "N/A"
                            valor_maior_servico = 0
                            percentual_maior_servico = 0
                            qtd_notas_maior_servico = 0
                            valor_medio_por_nota_maior_servico = 0
                            concentracao_top5_servicos = 0
                    else:
                        servico_maior_valor = "N/A"
                        valor_maior_servico = 0
                        percentual_maior_servico = 0
                        qtd_notas_maior_servico = 0
                        valor_medio_por_nota_maior_servico = 0
                        concentracao_top5_servicos = 0

                    indicadores_valor_servico = [
                        ("Serviço com Maior Valor", servico_maior_valor, "Código de serviço com maior valor total"),
                        ("Valor Total do Serviço", formatar_valor_br(valor_maior_servico), "Valor total do serviço com maior valor"),
                        ("Percentual do Serviço com Maior Valor", f"{percentual_maior_servico:.2f}%", "Percentual do serviço com maior valor em relação ao total"),
                        ("Quantidade de Notas do Serviço", qtd_notas_maior_servico, "Número de notas do serviço com maior valor"),
                        ("Valor Médio por Nota", formatar_valor_br(valor_medio_por_nota_maior_servico), "Valor médio por nota do serviço com maior valor"),
                        ("Concentração Top 5 Serviços", f"{concentracao_top5_servicos:.2f}%", "Percentual do faturamento dos 5 serviços com maior valor"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_valor_servico):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # 5.3 Tabela de Serviços
                    current_row += 1
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "5.3 Tabela de Serviços")

                    # Adiciona os top 5 serviços por valor
                    if "Item Lista Serviço" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        # Adiciona subtítulo para os top 5 serviços
                        current_row += 1
                        ws_indicadores.cell(row=current_row, column=1).value = "Top 5 Serviços por Valor"
                        ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                        ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                        ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                        current_row += 1

                        # Adiciona cabeçalhos para os top 5 serviços
                        ws_indicadores.cell(row=current_row, column=1).value = "Código de Serviço"
                        ws_indicadores.cell(row=current_row, column=2).value = "Valor Total"
                        ws_indicadores.cell(row=current_row, column=3).value = "% do Faturamento"

                        for col in range(1, 4):
                            ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                            ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                            ws_indicadores.cell(row=current_row, column=col).border = thin_border
                        current_row += 1

                        # Adiciona os top 5 serviços
                        top5_servicos = servicos_por_valor.head(5)

                        for i, (servico, valor) in enumerate(top5_servicos.items()):
                            percentual = (valor / total_base_calculo) * 100 if total_base_calculo > 0 else 0

                            ws_indicadores.cell(row=current_row, column=1).value = servico
                            ws_indicadores.cell(row=current_row, column=2).value = formatar_valor_br(valor)
                            ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                            # Formata células
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                            ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                            # Adiciona cor de fundo alternada para facilitar a leitura
                            if i % 2 == 0:
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                            # Adiciona borda
                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border

                            current_row += 1

                        # Adiciona subtítulo para os top 5 serviços por quantidade
                        current_row += 1
                        ws_indicadores.cell(row=current_row, column=1).value = "Top 5 Serviços por Quantidade de Notas"
                        ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                        ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                        ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                        current_row += 1

                        # Adiciona cabeçalhos para os top 5 serviços por quantidade
                        ws_indicadores.cell(row=current_row, column=1).value = "Código de Serviço"
                        ws_indicadores.cell(row=current_row, column=2).value = "Quantidade de Notas"
                        ws_indicadores.cell(row=current_row, column=3).value = "% do Total"

                        for col in range(1, 4):
                            ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                            ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                            ws_indicadores.cell(row=current_row, column=col).border = thin_border
                        current_row += 1

                        # Adiciona os top 5 serviços por quantidade
                        top5_servicos_qtd = servicos_mais_prestados.head(5)

                        for i, (servico, qtd) in enumerate(top5_servicos_qtd.items()):
                            percentual = (qtd / total_notas) * 100 if total_notas > 0 else 0

                            ws_indicadores.cell(row=current_row, column=1).value = servico
                            ws_indicadores.cell(row=current_row, column=2).value = qtd
                            ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                            # Formata células
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                            ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                            # Adiciona cor de fundo alternada para facilitar a leitura
                            if i % 2 == 0:
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                            # Adiciona borda
                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border

                            current_row += 1

                    # Adiciona espaço entre seções
                    current_row += 1

                    # Seção 6: Indicadores Geográficos
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "6. INDICADORES GEOGRÁFICOS")

                    # 6.1 Indicadores Gerais Geográficos
                    current_row = adicionar_titulo_secao(ws_indicadores, current_row, "6.1 Indicadores Gerais Geográficos")
                    current_row = adicionar_cabecalhos(ws_indicadores, current_row)

                    # Calcula o município com maior valor
                    if "Município" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        df_mun_valor = pd.DataFrame({
                            'Município': df_notas['Município'],
                            'Base de Cálculo': df_temp['Base de Cálculo']
                        })

                        # Agrupa por município e soma os valores
                        municipios_por_valor = df_mun_valor.groupby('Município')['Base de Cálculo'].sum().sort_values(ascending=False)

                        if not municipios_por_valor.empty:
                            municipio_maior_valor = municipios_por_valor.index[0]
                            valor_municipio_maior = municipios_por_valor.iloc[0]
                            percentual_municipio_maior = (valor_municipio_maior / total_base_calculo) * 100 if total_base_calculo > 0 else 0
                        else:
                            municipio_maior_valor = "N/A"
                            valor_municipio_maior = 0
                            percentual_municipio_maior = 0
                    else:
                        municipio_maior_valor = "N/A"
                        valor_municipio_maior = 0
                        percentual_municipio_maior = 0

                    # Calcula a UF com maior valor
                    if "UF" in df_notas.columns and "Base de Cálculo" in df_temp.columns:
                        df_uf_valor = pd.DataFrame({
                            'UF': df_notas['UF'],
                            'Base de Cálculo': df_temp['Base de Cálculo']
                        })

                        # Agrupa por UF e soma os valores
                        ufs_por_valor = df_uf_valor.groupby('UF')['Base de Cálculo'].sum().sort_values(ascending=False)

                        if not ufs_por_valor.empty:
                            uf_maior_valor = ufs_por_valor.index[0]
                            valor_uf_maior = ufs_por_valor.iloc[0]
                            percentual_uf_maior = (valor_uf_maior / total_base_calculo) * 100 if total_base_calculo > 0 else 0
                        else:
                            uf_maior_valor = "N/A"
                            valor_uf_maior = 0
                            percentual_uf_maior = 0
                    else:
                        uf_maior_valor = "N/A"
                        valor_uf_maior = 0
                        percentual_uf_maior = 0

                    indicadores_geograficos = [
                        ("Município Mais Frequente", municipio_mais_comum, "Município com maior número de notas"),
                        ("Quantidade de Notas no Município", qtd_municipio_mais_comum, "Número de notas no município mais frequente"),
                        ("Município com Maior Valor", municipio_maior_valor, "Município com maior valor total em notas"),
                        ("Valor Total do Município", formatar_valor_br(valor_municipio_maior), "Valor total das notas do município com maior valor"),
                        ("Percentual do Município", f"{percentual_municipio_maior:.2f}%", "Percentual do faturamento representado pelo município com maior valor"),
                        ("UF Mais Frequente", uf_mais_comum, "Estado com maior número de notas"),
                        ("Quantidade de Notas na UF", qtd_uf_mais_comum, "Número de notas na UF mais frequente"),
                        ("UF com Maior Valor", uf_maior_valor, "Estado com maior valor total em notas"),
                        ("Valor Total da UF", formatar_valor_br(valor_uf_maior), "Valor total das notas da UF com maior valor"),
                        ("Percentual da UF", f"{percentual_uf_maior:.2f}%", "Percentual do faturamento representado pela UF com maior valor"),
                    ]

                    for i, (indicador, valor, obs) in enumerate(indicadores_geograficos):
                        current_row = adicionar_indicador(ws_indicadores, current_row, indicador, valor, obs, i)

                    # Adiciona os top 5 municípios por quantidade
                    if "Município" in df_notas.columns:
                        # Adiciona subtítulo para os top 5 municípios
                        current_row += 1
                        ws_indicadores.cell(row=current_row, column=1).value = "Top 5 Municípios por Quantidade de Notas"
                        ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                        ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                        ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                        current_row += 1

                        # Adiciona cabeçalhos para os top 5 municípios
                        ws_indicadores.cell(row=current_row, column=1).value = "Município"
                        ws_indicadores.cell(row=current_row, column=2).value = "Quantidade de Notas"
                        ws_indicadores.cell(row=current_row, column=3).value = "% do Total"

                        for col in range(1, 4):
                            ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                            ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                            ws_indicadores.cell(row=current_row, column=col).border = thin_border
                        current_row += 1

                        # Adiciona os top 5 municípios
                        municipios = df_notas.groupby("Município").size().sort_values(ascending=False)
                        top5_municipios = municipios.head(5)

                        for i, (municipio, qtd) in enumerate(top5_municipios.items()):
                            percentual = (qtd / total_notas) * 100 if total_notas > 0 else 0

                            ws_indicadores.cell(row=current_row, column=1).value = municipio
                            ws_indicadores.cell(row=current_row, column=2).value = qtd
                            ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                            # Formata células
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                            ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                            # Adiciona cor de fundo alternada para facilitar a leitura
                            if i % 2 == 0:
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                            # Adiciona borda
                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border

                            current_row += 1

                        # Adiciona os top 5 municípios por valor
                        if "Base de Cálculo" in df_temp.columns:
                            # Adiciona subtítulo para os top 5 municípios por valor
                            current_row += 1
                            ws_indicadores.cell(row=current_row, column=1).value = "Top 5 Municípios por Valor"
                            ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                            ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            current_row += 1

                            # Adiciona cabeçalhos para os top 5 municípios por valor
                            ws_indicadores.cell(row=current_row, column=1).value = "Município"
                            ws_indicadores.cell(row=current_row, column=2).value = "Valor Total"
                            ws_indicadores.cell(row=current_row, column=3).value = "% do Faturamento"

                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                                ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                                ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border
                            current_row += 1

                            # Adiciona os top 5 municípios por valor
                            df_mun_valor = pd.DataFrame({
                                'Município': df_notas['Município'],
                                'Base de Cálculo': df_temp['Base de Cálculo']
                            })

                            # Agrupa por município e soma os valores
                            municipios_por_valor = df_mun_valor.groupby('Município')['Base de Cálculo'].sum().sort_values(ascending=False)
                            top5_municipios_valor = municipios_por_valor.head(5)

                            for i, (municipio, valor) in enumerate(top5_municipios_valor.items()):
                                percentual = (valor / total_base_calculo) * 100 if total_base_calculo > 0 else 0

                                ws_indicadores.cell(row=current_row, column=1).value = municipio
                                ws_indicadores.cell(row=current_row, column=2).value = formatar_valor_br(valor)
                                ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                                # Formata células
                                ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                                ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                                ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                                # Adiciona cor de fundo alternada para facilitar a leitura
                                if i % 2 == 0:
                                    for col in range(1, 4):
                                        ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                                # Adiciona borda
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).border = thin_border

                                current_row += 1

                    # 6.2 Distribuição por UF
                    if "UF" in df_notas.columns:
                        # Adiciona subtítulo para a seção
                        current_row += 1
                        current_row = adicionar_titulo_secao(ws_indicadores, current_row, "6.2 Distribuição por UF")

                        # Adiciona subtítulo para as UFs por quantidade
                        current_row += 1
                        ws_indicadores.cell(row=current_row, column=1).value = "UFs por Quantidade de Notas"
                        ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                        ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                        ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                        current_row += 1

                        # Adiciona cabeçalhos para as UFs por quantidade
                        ws_indicadores.cell(row=current_row, column=1).value = "UF"
                        ws_indicadores.cell(row=current_row, column=2).value = "Quantidade de Notas"
                        ws_indicadores.cell(row=current_row, column=3).value = "% do Total"

                        for col in range(1, 4):
                            ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                            ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                            ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                            ws_indicadores.cell(row=current_row, column=col).border = thin_border
                        current_row += 1

                        # Adiciona as UFs por quantidade
                        ufs = df_notas.groupby("UF").size().sort_values(ascending=False)

                        for i, (uf, qtd) in enumerate(ufs.items()):
                            percentual = (qtd / total_notas) * 100 if total_notas > 0 else 0

                            ws_indicadores.cell(row=current_row, column=1).value = uf
                            ws_indicadores.cell(row=current_row, column=2).value = qtd
                            ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                            # Formata células
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                            ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                            # Adiciona cor de fundo alternada para facilitar a leitura
                            if i % 2 == 0:
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                            # Adiciona borda
                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border

                            current_row += 1

                        # Adiciona as UFs por valor
                        if "Base de Cálculo" in df_temp.columns:
                            # Adiciona subtítulo para as UFs por valor
                            current_row += 1
                            ws_indicadores.cell(row=current_row, column=1).value = "UFs por Valor Total"
                            ws_indicadores.cell(row=current_row, column=1).font = Font(bold=True)
                            ws_indicadores.merge_cells(f'A{current_row}:C{current_row}')
                            ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                            current_row += 1

                            # Adiciona cabeçalhos para as UFs por valor
                            ws_indicadores.cell(row=current_row, column=1).value = "UF"
                            ws_indicadores.cell(row=current_row, column=2).value = "Valor Total"
                            ws_indicadores.cell(row=current_row, column=3).value = "% do Faturamento"

                            for col in range(1, 4):
                                ws_indicadores.cell(row=current_row, column=col).font = Font(bold=True)
                                ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                                ws_indicadores.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                                ws_indicadores.cell(row=current_row, column=col).border = thin_border
                            current_row += 1

                            # Cria DataFrame para UF e valor
                            df_uf_valor = pd.DataFrame({
                                'UF': df_notas['UF'],
                                'Base de Cálculo': df_temp['Base de Cálculo']
                            })

                            # Agrupa por UF e soma os valores
                            ufs_por_valor = df_uf_valor.groupby('UF')['Base de Cálculo'].sum().sort_values(ascending=False)

                            # Adiciona as UFs por valor
                            for i, (uf, valor) in enumerate(ufs_por_valor.items()):
                                percentual = (valor / total_base_calculo) * 100 if total_base_calculo > 0 else 0

                                ws_indicadores.cell(row=current_row, column=1).value = uf
                                ws_indicadores.cell(row=current_row, column=2).value = formatar_valor_br(valor)
                                ws_indicadores.cell(row=current_row, column=3).value = f"{percentual:.2f}%"

                                # Formata células
                                ws_indicadores.cell(row=current_row, column=1).alignment = Alignment(horizontal='left')
                                ws_indicadores.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
                                ws_indicadores.cell(row=current_row, column=3).alignment = Alignment(horizontal='center')

                                # Adiciona cor de fundo alternada para facilitar a leitura
                                if i % 2 == 0:
                                    for col in range(1, 4):
                                        ws_indicadores.cell(row=current_row, column=col).fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

                                # Adiciona borda
                                for col in range(1, 4):
                                    ws_indicadores.cell(row=current_row, column=col).border = thin_border

                                current_row += 1

                    # Salva as alterações
                    wb.save(arquivo_saida)
                    logging.info("Formatação adicional aplicada com sucesso")
                except Exception as e:
                    logging.warning(f"Não foi possível aplicar formatação adicional: {str(e)}")
                    logging.warning("O arquivo Excel foi salvo, mas sem formatação adicional")
            else:
                logging.info("Arquivo muito grande, pulando formatação adicional para economizar memória")
                print("Arquivo muito grande, pulando formatação adicional para economizar memória")

        except PermissionError:
            logging.error(f"Erro de permissão ao salvar o arquivo {arquivo_saida}. O arquivo pode estar aberto em outro programa.")
            print(f"ERRO: Não foi possível salvar o arquivo Excel. O arquivo pode estar aberto em outro programa.")
            print(f"Feche o arquivo Excel e tente novamente.")
            return
        except Exception as e:
            logging.error(f"Erro ao salvar o arquivo Excel: {str(e)}")
            print(f"ERRO: Não foi possível salvar o arquivo Excel: {str(e)}")
            return

        logging.info(f"Arquivo Excel gerado com sucesso: {arquivo_saida}")
        print(f"Processo concluído! {arquivos_processados} arquivos processados com sucesso.")
        if arquivos_com_erro > 0:
            print(f"Atenção: {arquivos_com_erro} arquivos não puderam ser processados. Verifique o log para mais detalhes.")
    else:
        logging.warning("Nenhum dado foi extraído dos arquivos XML")
        print("Nenhum dado foi extraído dos arquivos XML. Verifique se os arquivos estão no formato correto.")

if __name__ == "__main__":
    main()